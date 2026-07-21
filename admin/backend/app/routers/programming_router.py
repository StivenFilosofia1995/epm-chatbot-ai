import calendar
import io
import re
import unicodedata
from datetime import date
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from typing import Annotated
from ..deps import get_current_admin
from ..supabase_client import supabase
from ..schemas import ScheduleItem, ReplaceMonthInput
from ..pdf_ingestion import extract_text_from_pdf_bytes, parse_programming_text


router = APIRouter(prefix='/programming', tags=['programming'])

# Encabezados reales de los Excel que envía la Fundación EPM cada mes
# (ver "Programación infantil" / "Jóvenes y adultos" — una hoja por segmento,
# un archivo por espacio/UVA; el nombre del espacio NO viene como columna).
_EPM_COLUMNAS = {
    'titulo': ['titulo del curso', 'titulo', 'nombre del curso', 'curso', 'actividad'],
    'descripcion': ['descripcion'],
    'dias': ['dia(s)', 'dias', 'dia'],
    'fechas': ['fecha(s)', 'fechas', 'fecha'],
    'horario': ['horario', 'hora'],
    'lugar': ['lugar', 'espacio', 'sala'],
    'publico': ['publico', 'edad', 'rango de edad', 'edad_recomendada'],
    'inscripcion': ['inscripcion'],
    'enlace_inscripcion': ['enlace de inscripcion', 'enlace inscripcion', 'link de inscripcion', 'enlace'],
}

_MESES_ES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}

# Índice de weekday() de Python: lunes=0 ... domingo=6
_DIAS_SEMANA_ES = {
    'lunes': 0, 'martes': 1, 'miercoles': 2, 'jueves': 3, 'viernes': 4, 'sabado': 5, 'domingo': 6,
}


def _sin_tildes(texto: str) -> str:
    texto = unicodedata.normalize('NFKD', str(texto))
    return ''.join(ch for ch in texto if not unicodedata.combining(ch))


def _texto(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _mapear_encabezados(fila_encabezado: list, columnas: dict) -> dict[str, int]:
    """Retorna {campo_interno: indice_columna} según los encabezados detectados."""
    normalizados = [
        _sin_tildes(str(c).strip().lower()) if c is not None else ''
        for c in fila_encabezado
    ]
    mapeo: dict[str, int] = {}
    for campo, alias in columnas.items():
        for idx, encabezado in enumerate(normalizados):
            if encabezado in alias:
                mapeo[campo] = idx
                break
    return mapeo


def _expandir_fechas(texto_fecha: str, anio: int, mes: int) -> list[str]:
    """Convierte el texto libre de la columna 'Fecha(s)' en una lista de fechas ISO.

    Cubre los patrones reales observados en los Excel de EPM:
      - "01 de julio"                    → un solo día
      - "7, 14, 21 y 28 de julio"        → varios días explícitos
      - "23 y 30 de julio"               → idem, con solo dos
      - "Todos los martes de julio"      → recurrencia semanal en todo el mes
    El año y el mes SIEMPRE vienen del formulario de carga (el texto de la
    celda nunca incluye el año, y a veces ni el mes).
    """
    norm = _sin_tildes(texto_fecha).lower()

    dias_en_mes = calendar.monthrange(anio, mes)[1]
    fechas: list[str] = []

    m = re.search(r'todos\s+los\s+(lunes|martes|miercoles|jueves|viernes|sabado|domingo)', norm)
    if m:
        weekday_objetivo = _DIAS_SEMANA_ES[m.group(1)]
        for dia in range(1, dias_en_mes + 1):
            if date(anio, mes, dia).weekday() == weekday_objetivo:
                fechas.append(date(anio, mes, dia).isoformat())
        return fechas

    for dia_str in re.findall(r'\b([0-3]?\d)\b', norm):
        dia = int(dia_str)
        if 1 <= dia <= dias_en_mes:
            fechas.append(date(anio, mes, dia).isoformat())

    # Dedup preservando orden (por si el número del día se repite en el texto)
    vistas = set()
    unicas = []
    for f in fechas:
        if f not in vistas:
            vistas.add(f)
            unicas.append(f)
    return unicas


def _parsear_horario(texto_horario: str) -> tuple[str | None, str | None]:
    """Convierte "2:00 p.m. a 4:00 p.m." / "10:00 a.m. a 12:00 m." (mediodía) a 24h."""
    norm = texto_horario.lower()

    patron = re.compile(
        r'(\d{1,2}):(\d{2})\s*([ap]\.?\s*m\.?|m\.?)\s*a\s*(\d{1,2}):(\d{2})\s*([ap]\.?\s*m\.?|m\.?)',
        re.IGNORECASE,
    )
    match = patron.search(norm)
    if not match:
        return None, None

    h1, m1, mer1, h2, m2, mer2 = match.groups()
    return _hora_24(h1, m1, mer1), _hora_24(h2, m2, mer2)


def _hora_24(h: str, m: str, meridiano: str) -> str:
    hora = int(h)
    mer = re.sub(r'[.\s]', '', meridiano.lower())
    if mer == 'm':
        # "12:00 m." = mediodía (meridiano), no confundir con p.m.
        hora = 12
    elif mer == 'am':
        hora = 0 if hora == 12 else hora
    elif mer == 'pm':
        hora = hora if hora == 12 else hora + 12
    return f'{hora:02d}:{int(m):02d}'


def parse_programming_excel(file_bytes: bytes, uva_nombre: str, anio: int, mes: int) -> tuple[list[dict], dict]:
    """Parsea el Excel mensual real de EPM (columnas: Título del curso, Descripción,
    Día(s), Fecha(s), Horario, Lugar, Público, Inscripción, Enlace de inscripción).

    El archivo completo corresponde a UN espacio/UVA (no hay columna para eso),
    por lo que `uva_nombre` se recibe aparte y se aplica a todas las filas de
    todas las hojas (ej. "Programación infantil" + "Jóvenes y adultos").

    Cada fila puede expandirse a VARIAS filas en la base de datos si su columna
    'Fecha(s)' contiene más de una fecha o un patrón recurrente ("Todos los martes").
    """
    from openpyxl import load_workbook

    debug = {
        'hojas_leidas': 0, 'filas_totales': 0, 'filas_descartadas': 0,
        'fechas_generadas': 0, 'columnas_detectadas': {},
    }
    items: list[dict] = []

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    for hoja in wb.worksheets:
        filas = hoja.iter_rows(values_only=True)
        try:
            encabezado = next(filas)
        except StopIteration:
            continue

        mapeo = _mapear_encabezados(list(encabezado), _EPM_COLUMNAS)
        if 'titulo' not in mapeo or 'fechas' not in mapeo:
            continue  # hoja sin las columnas mínimas — se ignora (ej. una hoja de notas)

        debug['hojas_leidas'] += 1
        debug['columnas_detectadas'][hoja.title] = list(mapeo.keys())

        def _get(fila, campo):
            idx = mapeo.get(campo)
            return fila[idx] if idx is not None and idx < len(fila) else None

        for fila in filas:
            if fila is None or all(v is None or str(v).strip() == '' for v in fila):
                continue
            debug['filas_totales'] += 1

            titulo = _texto(_get(fila, 'titulo'))
            texto_fecha = _texto(_get(fila, 'fechas'))

            if not titulo or not texto_fecha:
                debug['filas_descartadas'] += 1
                continue

            fechas = _expandir_fechas(texto_fecha, anio, mes)
            if not fechas:
                debug['filas_descartadas'] += 1
                continue

            texto_horario = _texto(_get(fila, 'horario'))
            hora_inicio, hora_fin = _parsear_horario(texto_horario) if texto_horario else (None, None)

            partes_desc = []
            descripcion = _texto(_get(fila, 'descripcion'))
            if descripcion:
                partes_desc.append(descripcion)
            lugar = _texto(_get(fila, 'lugar'))
            if lugar:
                partes_desc.append(f'Lugar: {lugar}')
            inscripcion = _texto(_get(fila, 'inscripcion'))
            if inscripcion and 'no requiere' not in _sin_tildes(inscripcion).lower():
                partes_desc.append(f'Inscripción: {inscripcion}')
            enlace = _texto(_get(fila, 'enlace_inscripcion'))
            if enlace:
                partes_desc.append(f'Enlace de inscripción: {enlace}')
            descripcion_final = ' | '.join(partes_desc)[:500] or None

            edad_recomendada = _texto(_get(fila, 'publico'))

            for fecha_iso in fechas:
                items.append({
                    'uva_nombre': uva_nombre,
                    'fecha': fecha_iso,
                    'hora_inicio': hora_inicio,
                    'hora_fin': hora_fin,
                    'actividad': titulo[:240],
                    'descripcion': descripcion_final,
                    'edad_recomendada': edad_recomendada,
                })
                debug['fechas_generadas'] += 1

    return items, debug


def _known_uvas() -> list[str]:
    rows = (
        supabase
        .table('programacion_uva')
        .select('uva_nombre')
        .limit(5000)
        .execute()
    ).data or []
    found = {row.get('uva_nombre', '').strip() for row in rows if row.get('uva_nombre')}
    if not found:
        found = {
            'UVA de La Imaginacion',
            'UVA El Paraiso',
            'UVA Nuevo Occidente',
            'UVA Sol de Oriente',
            'UVA Mirador de Calasanz',
        }
    return sorted(found)


def _replace_month_if_requested(first_date: str | None):
    if not first_date:
        return
    year, month, _ = first_date.split('-')
    first_day = f'{int(year):04d}-{int(month):02d}-01'
    next_month = f'{int(year) + (1 if int(month) == 12 else 0):04d}-{(1 if int(month) == 12 else int(month) + 1):02d}-01'
    supabase.table('programacion_uva').delete().gte('fecha', first_day).lt('fecha', next_month).execute()


@router.get('')
def list_programming(
    _admin: Annotated[str, Depends(get_current_admin)],
    fecha: Annotated[str | None, Query()] = None,
    uva: Annotated[str | None, Query()] = None,
    limit: Annotated[int, Query(ge=1, le=5000)] = 500,
):
    q = supabase.table('programacion_uva').select('*').order('fecha').order('hora_inicio').limit(limit)
    if fecha:
        q = q.eq('fecha', fecha)
    if uva:
        q = q.eq('uva_nombre', uva)
    result = q.execute()
    return {'items': result.data or []}


@router.post('/upsert')
def upsert_programming(item: ScheduleItem, _admin: Annotated[str, Depends(get_current_admin)]):
    payload = item.model_dump()
    supabase.table('programacion_uva').insert(payload).execute()
    return {'ok': True, 'item': payload}


@router.delete('/by-month')
def delete_month(year: int, month: int, _admin: Annotated[str, Depends(get_current_admin)]):
    first_day = f'{year:04d}-{month:02d}-01'
    next_month = f'{year + (1 if month == 12 else 0):04d}-{(1 if month == 12 else month + 1):02d}-01'
    supabase.table('programacion_uva').delete().gte('fecha', first_day).lt('fecha', next_month).execute()
    return {'ok': True, 'deleted_range': [first_day, next_month]}


@router.post('/replace-month')
def replace_month(data: ReplaceMonthInput, _admin: Annotated[str, Depends(get_current_admin)]):
    first_day = f'{data.year:04d}-{data.month:02d}-01'
    next_month = f'{data.year + (1 if data.month == 12 else 0):04d}-{(1 if data.month == 12 else data.month + 1):02d}-01'
    supabase.table('programacion_uva').delete().gte('fecha', first_day).lt('fecha', next_month).execute()
    return {'ok': True, 'month': data.month, 'year': data.year}


@router.post('/ingest-pdf')
async def ingest_pdf_programming(
    _admin: Annotated[str, Depends(get_current_admin)],
    file: UploadFile = File(...),
    replace_month: bool = Form(False),
    ocr_lang: str = Form('spa'),
):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail='Solo se permiten archivos PDF')

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail='El PDF esta vacio')

    # CRITICO: extract_text_from_pdf_bytes (OCR, puede tardar decenas de
    # segundos) y las llamadas a Supabase son SINCRONAS/bloqueantes. Este
    # endpoint es `async def` (necesario por `await file.read()`), y FastAPI
    # NO offloadea automaticamente el cuerpo de un handler async a un thread
    # — a diferencia de un handler sync normal. Sin run_in_threadpool, una
    # llamada bloqueante aca congela el ÚNICO event loop de uvicorn para
    # TODA la app: cualquier otra peticion (incluso a rutas totalmente
    # distintas, como reset-total) queda colgada hasta que esta termine.
    text, extract_debug = await run_in_threadpool(extract_text_from_pdf_bytes, pdf_bytes, ocr_lang=ocr_lang)
    if not text.strip():
        raise HTTPException(status_code=422, detail='No se pudo extraer texto util del PDF')

    parsed = parse_programming_text(text, _known_uvas())
    if not parsed.items:
        raise HTTPException(
            status_code=422,
            detail='No se detectaron actividades. Verifique formato del PDF o active OCR en el servidor.',
        )

    if replace_month:
        await run_in_threadpool(_replace_month_if_requested, parsed.items[0].get('fecha'))

    payload = [
        {
            'uva_nombre': item['uva_nombre'],
            'fecha': item['fecha'] or date.today().isoformat(),
            'hora_inicio': item.get('hora_inicio'),
            'hora_fin': item.get('hora_fin'),
            'actividad': item.get('actividad') or 'Actividad por confirmar',
            'descripcion': item.get('descripcion'),
            'edad_recomendada': item.get('edad_recomendada'),
        }
        for item in parsed.items
    ]

    await run_in_threadpool(lambda: supabase.table('programacion_uva').insert(payload).execute())

    return {
        'ok': True,
        'archivo': file.filename,
        'insertados': len(payload),
        'extract_debug': extract_debug,
        'parse_debug': parsed.debug,
    }


@router.post('/ingest-excel')
async def ingest_excel_programming(
    _admin: Annotated[str, Depends(get_current_admin)],
    file: UploadFile = File(...),
    uva_nombre: str = Form(...),
    anio: int = Form(...),
    mes: int = Form(..., ge=1, le=12),
    replace_month: bool = Form(False),
):
    """Carga la programación mensual desde el Excel real que envía EPM (.xlsx).

    El archivo corresponde a UN espacio/UVA completo (Biblioteca EPM, una UVA, etc.),
    con hojas como "Programación infantil" / "Jóvenes y adultos". Columnas esperadas
    en cada hoja (encabezado en la primera fila, cualquier orden): Título del curso,
    Descripción, Día(s), Fecha(s), Horario, Lugar, Público, Inscripción,
    Enlace de inscripción. El nombre del espacio y el mes/año se indican aparte
    porque el archivo no los trae como columna.

    Es la vía recomendada para actualizar el mes: no depende de scraping ni OCR.
    """
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail='Solo se permiten archivos Excel (.xlsx)')

    excel_bytes = await file.read()
    if not excel_bytes:
        raise HTTPException(status_code=400, detail='El archivo esta vacio')

    try:
        # parse_programming_excel + las llamadas a Supabase son sincronas —
        # ver la nota extensa en ingest_pdf_programming: sin run_in_threadpool
        # aca, esto bloquea el event loop entero de uvicorn para TODA la app.
        items, parse_debug = await run_in_threadpool(
            parse_programming_excel, excel_bytes, uva_nombre.strip(), anio, mes,
        )
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f'No se pudo leer el Excel: {exc}')

    if not items:
        raise HTTPException(
            status_code=422,
            detail='No se detectaron filas válidas. Verifique que la primera fila tenga los '
                   'encabezados Título del curso, Fecha(s), Horario, etc.',
        )

    if replace_month:
        # Acotado a este espacio: cada Excel cubre UN solo espacio/UVA, así que
        # reemplazar el mes NO debe borrar la programación de los demás espacios
        # (a diferencia de /replace-month y el PDF, que cubren todos a la vez).
        first_day = f'{anio:04d}-{mes:02d}-01'
        next_month = f'{anio + (1 if mes == 12 else 0):04d}-{(1 if mes == 12 else mes + 1):02d}-01'

        def _borrar_mes_espacio():
            supabase.table('programacion_uva') \
                .delete() \
                .eq('uva_nombre', uva_nombre.strip()) \
                .gte('fecha', first_day) \
                .lt('fecha', next_month) \
                .execute()

        await run_in_threadpool(_borrar_mes_espacio)

    await run_in_threadpool(lambda: supabase.table('programacion_uva').insert(items).execute())

    return {
        'ok': True,
        'archivo': file.filename,
        'uva_nombre': uva_nombre,
        'insertados': len(items),
        'parse_debug': parse_debug,
    }


@router.get('/coverage')
def programming_coverage(_admin: Annotated[str, Depends(get_current_admin)]):
    """Resume qué tan al día está la programación cargada.

    Se usa en el panel para advertir cuando el mes actual no tiene datos —
    la causa más común de que el bot responda "no tengo programación" seguido.
    """
    rows = (
        supabase
        .table('programacion_uva')
        .select('fecha')
        .order('fecha')
        .limit(20000)
        .execute()
    ).data or []

    fechas = sorted({r['fecha'] for r in rows if r.get('fecha')})
    hoy = date.today().isoformat()
    mes_actual = hoy[:7]
    tiene_mes_actual = any(f.startswith(mes_actual) for f in fechas)

    return {
        'ok': True,
        'hoy': hoy,
        'primera_fecha': fechas[0] if fechas else None,
        'ultima_fecha': fechas[-1] if fechas else None,
        'total_fechas_distintas': len(fechas),
        'mes_actual_cubierto': tiene_mes_actual,
        'advertencia': None if tiene_mes_actual else (
            f'No hay programación cargada para el mes actual ({mes_actual}). '
            'El bot responderá "no tengo programación" a la mayoría de consultas '
            'hasta que se cargue el Excel o PDF de este mes.'
        ),
    }
